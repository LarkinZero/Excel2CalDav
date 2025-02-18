from openpyxl import load_workbook
from datetime import datetime, timedelta
import json
import config

year = 2025


def get_cell_value(sheet, cell):
    """
    获取单元格的值，如果单元格在合并单元格范围内，返回合并单元格的值
    """
    # 检查指定单元格是否在合并单元格范围内
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # 如果单元格在合并单元格范围内，返回合并单元格的值
            return sheet[merged_range.start_cell.coordinate].value
    return cell.value


def find_cell_with_value(sheet, value):
    """
    查找包含指定值的单元格
    """
    result = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == value:
                result.append(cell)
    return result


def format_date(date_str: str) -> datetime:
    """
    格式化日期字符串
    :param date_str: 日期字符串
    :return: 日期对象
    """
    format1 = "%m.%d"
    format2 = "%m月%d日"
    if "月" in date_str:
        date = datetime.strptime(date_str, format2).replace(year=year)
    else:
        date = datetime.strptime(date_str, format1).replace(year=year)
    return date


def generate_calendar_workdays(cal: dict, workbook_path: str) -> dict:
    """
    生成工作日历
    :param cal: 日历字典
    :param workbook_path: 工作簿路径
    :return: 日历字典
    ''
    """
    wb = load_workbook(workbook_path)
    sheet = wb.active
    result = find_cell_with_value(sheet, config.vip_name)
    for i in result:
        header = get_cell_value(sheet, sheet[i.column_letter + '3']).replace('\n', '')
        date = sheet['A' + str(i.row)].value
        cal[format_date(date).isoformat()] = header
    return cal


def generate_calendar_holidays(cal: dict, workbook_path: str) -> dict:
    """
    生成节假日日历
    :param cal: 日历字典
    :param workbook_path: 工作簿路径
    :return: 日历字典
    ''
    """
    wb2 = load_workbook(workbook_path)
    sheet2 = wb2.active
    result2 = find_cell_with_value(sheet2, config.vip_name)
    for i in result2:
        # header = get_cell_value(sheet2,sheet2["A"+str(i.row)])
        header = '值班'
        date = sheet2[i.column_letter + '2'].value
        cal[format_date(date).isoformat()] = header
    return cal


def date_format(date_str: str) -> datetime:
    return datetime.strptime(date_str, '%Y-%m-%dT%H:%M:%S')


def make_month_dict(date: str):
    start_date = date_format(date)
    first_day_of_month = start_date.replace(day=1)
    if first_day_of_month.month in range(0, 12):
        next_month = first_day_of_month.replace(month=first_day_of_month.month + 1)
    else:
        next_month = first_day_of_month.replace(year=first_day_of_month.year + 1, month=1)
    last_day_of_month = next_month - timedelta(days=1)
    month_dict = []
    current_date = first_day_of_month
    while current_date <= last_day_of_month:
        month_dict.append(current_date.strftime('%Y-%m-%dT%H:%M:%S'))
        current_date += timedelta(days=1)
    return month_dict


def filter_calendar_by_month(cal: dict, month: int) -> dict:
    """
    从日历字典中筛选出指定月份的条目。

    :param cal: 包含日期和事件的字典。
    :param month: 指定的月份。
    :return: 包含指定月份条目的新字典。
    """
    filtered_items = filter(
        lambda item: date_format(item[0]).month == month,
        cal.items()
    )

    # 将筛选结果转换为字典
    return dict(filtered_items)


def isWeekend(date_str: str) -> bool:
    return date_format(date_str).weekday() > 4


def makeWeekend(cal: dict) -> dict:
    first_date = list(cal.keys())[0]
    month_dict = make_month_dict(first_date)
    for i in month_dict:
        if i not in cal and isWeekend(i):
            cal[i] = '休息'
    return cal


if __name__ == '__main__':
    calendar = {}
    calendar = generate_calendar_workdays(calendar, r'./example/25年2月服务窗口排班表.xlsx')
    calendar = generate_calendar_holidays(calendar, r'./example/25年2月节假日排班表.xlsx')
    calendar = filter_calendar_by_month(calendar, 2)
    calendar = makeWeekend(calendar)
    print(json.dumps(calendar, indent=4, ensure_ascii=False))
